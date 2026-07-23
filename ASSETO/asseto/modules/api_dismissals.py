"""Dismissals API Blueprint — employee offboarding workflow."""

import os, io, uuid, json, hashlib, hmac as _hmac
from datetime import date, datetime

from flask import Blueprint, render_template, request, jsonify, send_file, abort

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from modules.config import app, ROLES, CONDITIONS, UPLOADS, SIGS
from modules.db import get_db
from modules.auth import (login_required, roles_required, get_current_user,
                          _save_signature, bhost)
from modules.api_items import log_h

bp = Blueprint('dismissals', __name__)


# ══════════════════════════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _dismissal_token(dis):
    return hashlib.sha256(
        f"{dis.get('employee_email','')}{dis['id']}{app.config['SECRET_KEY']}".encode()
    ).hexdigest()[:16]


def send_tg_notification(user_id, message):
    """Helper to send Telegram notifications if chat_id exists."""
    with get_db() as db:
        user = db.execute("SELECT telegram_chat_id FROM users WHERE id=?", (user_id,)).fetchone()
    if user and user["telegram_chat_id"]:
        token = os.environ.get('TELEGRAM_BOT_TOKEN')
        if not token: return
        try:
            import urllib.request as _ur
            url = f"https://api.telegram.org/bot{token}/sendMessage"
            data = {"chat_id": user["telegram_chat_id"], "text": message, "parse_mode": "HTML"}
            req = _ur.Request(url, data=json.dumps(data).encode(), headers={'Content-Type': 'application/json'})
            with _ur.urlopen(req) as resp:
                pass
        except Exception as e:
            print(f"  [!] Telegram error: {e}")


# ══════════════════════════════════════════════════════════════════════════════
#  DISMISSAL WORKFLOW — HTML PAGES
# ══════════════════════════════════════════════════════════════════════════════

@bp.route("/admin/dismissals")
@roles_required("superadmin","aho","hr")
def admin_dismissals_page():
    with get_db() as db:
        rows = db.execute("SELECT * FROM dismissals ORDER BY created_at DESC").fetchall()
    return render_template("admin_dismissals.html", dismissals=[dict(r) for r in rows],
                           user=request.current_user)


@bp.route("/dismissal/<int:did>")
def dismissal_page(did):
    """Public page for employee to submit photos — guarded by employee email token."""
    token = request.args.get("t", "")
    with get_db() as db:
        dis = db.execute("SELECT * FROM dismissals WHERE id=?", (did,)).fetchone()
    if not dis:
        abort(404)
    dis = dict(dis)
    # Guard: require either a valid session OR the employee's email hash as ?t= token
    u = get_current_user()
    if not u:
        expected = hashlib.sha256(
            f"{dis['employee_email']}{dis['id']}{app.config['SECRET_KEY']}".encode()
        ).hexdigest()[:16]
        if not token or token != expected:
            return render_template("login.html"), 403
    dis["items"] = json.loads(dis["items_json"] or "[]")
    dis["photos"] = json.loads(dis["photos_json"] or "{}")
    return render_template("dismissal.html", dis=dis)


# ══════════════════════════════════════════════════════════════════════════════
#  DISMISSAL WORKFLOW — API
# ══════════════════════════════════════════════════════════════════════════════

@bp.route("/api/dismissals", methods=["POST"])
@roles_required("superadmin","aho","hr")
def create_dismissal():
    d = request.json or {}
    u = request.current_user
    emp_id = d.get("employee_id")
    if not emp_id:
        return jsonify({"error": "Не указан сотрудник"}), 400
    with get_db() as db:
        # Resolve employee by ID or Name
        emp = db.execute("SELECT * FROM users WHERE id=? OR name=?", (emp_id, emp_id)).fetchone()
        if not emp:
            return jsonify({"error": "Сотрудник не найден"}), 404
        emp_id = emp["id"]

        # Get all items assigned to employee
        items = db.execute(
            "SELECT id, inv_num, category, model, room FROM items WHERE employee_id=? OR employee=?",
            (emp_id, emp["name"])
        ).fetchall()
        items_json = json.dumps([dict(i) for i in items])
        # Check if active dismissal already exists
        existing = db.execute(
            "SELECT id FROM dismissals WHERE employee_id=? AND status NOT IN ('completed','cancelled')",
            (emp_id,)
        ).fetchone()
        if existing:
            return jsonify({"error": "Процесс увольнения уже запущен", "id": existing["id"]}), 400
        cur = db.execute(
            "INSERT INTO dismissals (employee_id,employee_name,employee_email,initiated_by,initiated_by_name,items_json,notes,status) VALUES (?,?,?,?,?,?,?,'pending_aho')",
            (emp_id, emp["name"], emp["email"], u["id"], u["name"], items_json, d.get("notes",""))
        )
        dis_id = cur.lastrowid
    token = hashlib.sha256(
        f"{emp['email']}{dis_id}{app.config['SECRET_KEY']}".encode()
    ).hexdigest()[:16]
    return jsonify({"ok": True, "dismissal_id": dis_id, "access_token": token})


@bp.route("/api/dismissals/<int:did>/aho_accept", methods=["POST"])
@roles_required("superadmin","aho")
def dismissals_aho_accept(did):
    d = request.json or {}
    u = request.current_user
    with get_db() as db:
        dis = db.execute("SELECT * FROM dismissals WHERE id=?", (did,)).fetchone()
        if not dis: return jsonify({"error": "Не найдено"}), 404
        dis = dict(dis)
        if dis["status"] not in ("pending_aho", "pending", "photos_submitted"):
            return jsonify({"error": "Неверный статус"}), 400

        # Return all items
        items = json.loads(dis["items_json"])
        cond_map = json.loads(dis.get("item_conditions") or "{}")
        for item in items:
            iid = item["id"] if isinstance(item, dict) else item
            cond = cond_map.get(str(iid)) or d.get(str(iid)) or "Хорошее"
            if cond not in CONDITIONS: cond = "Хорошее"
            if cond == "Утеряно":
                db.execute("UPDATE items SET condition='Утеряно',employee_id=NULL,employee='—',status='Свободно' WHERE id=?",(iid,))
                log_h(db,iid,f"❌ Утеряно при увольнении: {dis['employee_name']}",uid=u["id"],uname=u["name"])
            else:
                db.execute("UPDATE items SET condition=?,employee_id=NULL,employee='—',status='Свободно' WHERE id=?",(cond,iid))
                log_h(db,iid,f"Возвращено (увольнение: {dis['employee_name']})",uid=u["id"],uname=u["name"])

        # Handle Signature
        sig_path = None
        if d.get("signature"):
            sig_path = _save_signature(d["signature"], f"dis_{did}_aho")

        db.execute(
            """UPDATE dismissals SET 
               status='pending_it', 
               aho_cleared=1, aho_at=CURRENT_TIMESTAMP, 
               aho_by_id=?, aho_by_name=?, aho_signature=?
               WHERE id=?""",
            (u["id"], u["name"], sig_path, did)
        )
    return jsonify({"ok": True})


@bp.route("/api/dismissals/<int:did>/it_accept", methods=["POST"])
@roles_required("superadmin","aho","deputy")
def dismissals_it_accept(did):
    d = request.json or {}
    u = request.current_user
    with get_db() as db:
        dis = db.execute("SELECT * FROM dismissals WHERE id=?", (did,)).fetchone()
        if not dis: return jsonify({"error": "Не найдено"}), 404
        dis = dict(dis)
        if dis["status"] != "pending_it":
            return jsonify({"error": "Сначала техника должна быть принята АХО"}), 400

        sig_path = None
        if d.get("signature"):
            sig_path = _save_signature(d["signature"], f"dis_{did}_it")

        db.execute(
            """UPDATE dismissals SET 
               status='pending_hr', 
               it_cleared=1, it_at=CURRENT_TIMESTAMP, 
               it_by_id=?, it_by_name=?, it_signature=?
               WHERE id=?""",
            (u["id"], u["name"], sig_path, did)
        )
    return jsonify({"ok": True})


@bp.route("/api/dismissals/<int:did>/hr_finalize", methods=["POST"])
@roles_required("superadmin","hr")
def dismissals_hr_finalize(did):
    u = request.current_user
    with get_db() as db:
        dis = db.execute("SELECT * FROM dismissals WHERE id=?", (did,)).fetchone()
        if not dis: return jsonify({"error": "Не найдено"}), 404
        dis = dict(dis)
        if dis["status"] != "pending_hr":
            return jsonify({"error": "Техника еще не принята АХО"}), 400

        # Handle Signature
        d = request.json or {}
        sig_path = None
        if d.get("signature"):
            sig_path = _save_signature(d["signature"], f"dis_{did}_hr")

        # Deactivate user
        db.execute("UPDATE users SET active=0 WHERE id=?", (dis["employee_id"],))
        db.execute(
            """UPDATE dismissals SET 
               status='completed', completed_at=CURRENT_TIMESTAMP,
               hr_at=CURRENT_TIMESTAMP, hr_by_id=?, hr_by_name=?, hr_signature=?
               WHERE id=?""",
            (u["id"], u["name"], sig_path, did)
        )
    return jsonify({"ok": True})


@bp.route("/api/dismissals/<int:did>/submit", methods=["POST"])
def submit_dismissal_photos(did):
    """Employee submits photos — guarded by session OR email token."""
    with get_db() as db:
        dis = db.execute("SELECT * FROM dismissals WHERE id=?", (did,)).fetchone()
    if not dis:
        return jsonify({"error": "Не найдено"}), 404
    dis = dict(dis)
    current_u = get_current_user()
    if not current_u:
        token = request.args.get("t") or request.form.get("_token", "")
        expected = hashlib.sha256(
            f"{dis.get('employee_email','')}{dis['id']}{app.config['SECRET_KEY']}".encode()
        ).hexdigest()[:16]
        if not token or not _hmac.compare_digest(token, expected):
            return jsonify({"error": "Недействительная ссылка"}), 403
    if dis["status"] not in ("pending", "photos_requested"):
        return jsonify({"error": "Форма уже отправлена"}), 400
    photos = {}
    for key in request.files:
        if key.startswith("photo_"):
            item_id = key.replace("photo_", "")
            f = request.files[key]
            ext = os.path.splitext(f.filename)[1].lower() or ".jpg"
            name = f"dis_{did}_{item_id}_{uuid.uuid4().hex[:6]}{ext}"
            f.save(os.path.join(UPLOADS, name))
            photos[item_id] = name
    try:
        conditions = json.loads(request.form.get("conditions", "{}") or "{}")
    except:
        conditions = {}

    # Save Employee Signature
    emp_sig_path = None
    if request.form.get("signature"):
        emp_sig_path = _save_signature(request.form["signature"], f"dis_{did}_emp")

    photos_data = {"photos": photos, "conditions": conditions,
                   "comment": request.form.get("comment", ""),
                   "submitted_at": datetime.now().isoformat()}
    with get_db() as db:
        db.execute("UPDATE dismissals SET status='photos_submitted', photos_json=?, employee_signature=? WHERE id=?",
                   (json.dumps(photos_data), emp_sig_path, did))
    return jsonify({"ok": True})


@bp.route("/api/dismissals/<int:did>")
@login_required
def get_dismissal(did):
    with get_db() as db:
        dis = db.execute("SELECT * FROM dismissals WHERE id=?", (did,)).fetchone()
    if not dis:
        return jsonify({"error": "Не найдено"}), 404
    dis = dict(dis)
    dis["items"] = json.loads(dis["items_json"] or "[]")
    dis["photos_data"] = json.loads(dis["photos_json"] or "{}")
    dis["access_token"] = _dismissal_token(dis)
    return jsonify(dis)


@bp.route("/api/dismissals/<int:did>/cancel", methods=["POST"])
@roles_required("superadmin","aho","hr")
def cancel_dismissal(did):
    with get_db() as db:
        db.execute("UPDATE dismissals SET status='cancelled' WHERE id=?", (did,))
    return jsonify({"ok": True})


@bp.route("/api/dismissals/<int:did>/request-photos", methods=["POST"])
@roles_required("superadmin","aho","hr")
def request_dismissal_photos(did):
    with get_db() as db:
        db.execute("UPDATE dismissals SET status='photos_requested' WHERE id=?", (did,))
    return jsonify({"ok": True})


@bp.route("/api/dismissals/<int:did>/receive-item", methods=["POST"])
@roles_required("superadmin", "aho")
def receive_dismissal_item(did):
    """Marks a specific item in the dismissal list as physically received."""
    d = request.json
    iid = d.get("item_id")
    if not iid: return jsonify({"error": "ID предмета не указан"}), 400

    with get_db() as db:
        dis = db.execute("SELECT items_json FROM dismissals WHERE id=?", (did,)).fetchone()
        if not dis: return jsonify({"error": "Обходной лист не найден"}), 404

        items = json.loads(dis["items_json"])
        found = False
        for item in items:
            if item["id"] == iid:
                item["received"] = True
                item["received_at"] = datetime.now().isoformat()
                item["received_by"] = request.current_user["name"]
                found = True
                break

        if not found: return jsonify({"error": "Предмет не найден в списке"}), 404

        db.execute("UPDATE dismissals SET items_json=? WHERE id=?", (json.dumps(items), did))
    return jsonify({"ok": True})


@bp.route("/api/dismissals/<int:did>/finalize", methods=["POST"])
@roles_required("superadmin", "hr")
def finalize_dismissal(did):
    """
    Finalizes the dismissal process ONLY if all items are received.
    """
    u = request.current_user
    with get_db() as db:
        dis = db.execute("SELECT * FROM dismissals WHERE id=?", (did,)).fetchone()
        if not dis: return jsonify({"error": "Обходной лист не найден"}), 404
        if dis["status"] == "completed": return jsonify({"error": "Уже завершено"}), 400

        # Check if all items are received
        items = json.loads(dis["items_json"])
        for item in items:
            if not item.get("received"):
                return jsonify({"error": f"Сначала примите все вещи (не принят: {item.get('inv_num')})"}), 400

        emp_id = dis["employee_id"]
        emp_name = dis["employee_name"]

        # Deactivate User and Release Assets
        db.execute("UPDATE users SET active=0 WHERE id=?", (emp_id,))
        asset_items = db.execute("SELECT id FROM items WHERE employee_id=? OR employee=?", (emp_id, emp_name)).fetchall()
        for item in asset_items:
            db.execute("UPDATE items SET employee_id=NULL, employee='—', status='Свободно' WHERE id=?", (item["id"],))
            log_h(db, item["id"], f"Освобождено (увольнение {emp_name})", uid=u["id"], uname=u["name"])

        db.execute("UPDATE dismissals SET status='completed', completed_at=CURRENT_TIMESTAMP, confirmed_by=?, confirmed_by_name=? WHERE id=?",
                   (u["id"], u["name"], did))

        send_tg_notification(dis["initiated_by"], f"<b>✅ Обходной лист закрыт</b>\nСотрудник: {emp_name}\nВсе вещи приняты АХО.")

    return jsonify({"ok": True})


# ══════════════════════════════════════════════════════════════════════════════
#  DISMISSAL EXPORT
# ══════════════════════════════════════════════════════════════════════════════

@bp.route("/api/dismissals/<int:did>/export")
@roles_required("superadmin","aho","hr","accountant")
def export_dismissal_act(did):
    with get_db() as db:
        dis = db.execute("SELECT * FROM dismissals WHERE id=?",(did,)).fetchone()
    if not dis: abort(404)
    dis = dict(dis)
    items_l = json.loads(dis.get("items_json") or "[]")
    cmap    = json.loads(dis.get("item_conditions") or "{}")
    commmap = json.loads(dis.get("item_comments") or "{}")
    wb=Workbook(); ws=wb.active; ws.title="Акт"
    blue="1E40AF"; thin=Side(style="thin",color="CBD5E1")
    bdr=Border(left=thin,right=thin,top=thin,bottom=thin)
    ws.merge_cells("A1:G1")
    ws["A1"]="АКТ ПРИЁМА-ПЕРЕДАЧИ МАТЕРИАЛЬНЫХ ЦЕННОСТЕЙ"
    ws["A1"].font=Font(bold=True,size=13,color="FFFFFF")
    ws["A1"].fill=PatternFill("solid",fgColor=blue)
    ws["A1"].alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height=30
    for i,(k,v) in enumerate([
        ("Сотрудник:",dis.get("employee_name","")),
        ("Дата:",date.today().strftime("%d.%m.%Y")),
        ("Инициировал:",dis.get("initiated_by_name",""))],3):
        ws[f"A{i}"]=k; ws[f"A{i}"].font=Font(bold=True)
        ws[f"B{i}"]=v; ws.merge_cells(f"B{i}:G{i}")
    hr=7
    for col,h in enumerate(["Инв.№","Категория","Модель","Кабинет","При выдаче","При возврате","Комментарий"],1):
        c=ws.cell(hr,col,h); c.font=Font(bold=True,color="FFFFFF")
        c.fill=PatternFill("solid",fgColor=blue); c.border=bdr
        c.alignment=Alignment(horizontal="center")
    for ri,item in enumerate(items_l,hr+1):
        iid=str(item.get("id",""))
        cr=cmap.get(iid,"—")
        fg="FEE2E2" if cr=="Утеряно" else ("F8FAFC" if ri%2==0 else "FFFFFF")
        for col,val in enumerate([item.get("inv_num",""),item.get("category",""),
            item.get("model",""),item.get("room",""),"Хорошее",cr,commmap.get(iid,"")],1):
            c=ws.cell(ri,col,val); c.border=bdr
            c.fill=PatternFill("solid",fgColor=fg)
    sr=hr+len(items_l)+2
    ws[f"A{sr}"]="Подпись АХО:"; ws[f"A{sr}"].font=Font(bold=True)
    ws[f"C{sr}"]="____________________"
    ws[f"E{sr}"]="Дата:"; ws[f"F{sr}"]=date.today().strftime("%d.%m.%Y")
    for i,w in enumerate([12,14,22,16,16,16,28],1):
        ws.column_dimensions[get_column_letter(i)].width=w
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    fn=f"Акт_{(dis.get('employee_name') or 'emp').replace(' ','_')}_{date.today()}.xlsx"
    return send_file(buf,as_attachment=True,download_name=fn,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
