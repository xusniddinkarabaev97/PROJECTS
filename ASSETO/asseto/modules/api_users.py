"""Users API Blueprint — user management, staff, departments, history, and related pages."""

import os, io, json, secrets, time
from datetime import date, datetime, timedelta

from flask import (Blueprint, render_template, request, jsonify, send_file,
                   abort, redirect, url_for, make_response)

import bcrypt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from modules.config import (app, ROLES, CATEGORIES, CONDITIONS, STATUSES,
                            UPLOADS, SIGS, SECURE_COOKIES, MAX_UPLOAD_MB,
                            JWT_EXPIRY, IntegrityError)
from modules.db import get_db, _trunc
from modules.auth import login_required, roles_required, bhost, rate_limit
from modules.api_items import qr_png, _db_error

bp = Blueprint('users', __name__)


# ══════════════════════════════════════════════════════════════════════════════
#  USERS CRUD API
# ══════════════════════════════════════════════════════════════════════════════

@bp.route("/api/users")
@roles_required("superadmin", "aho", "hr", "director")
def get_users():
    with get_db() as db:
        rows = db.execute("""
            SELECT u.id, u.name, u.email, u.role, u.department, u.active, u.created_at,
                   COUNT(i.id) as items_count
            FROM users u
            LEFT JOIN items i ON i.employee_id = u.id OR i.employee = u.name
            GROUP BY u.id
            ORDER BY u.name
        """).fetchall()
    return jsonify([dict(r) for r in rows])


@bp.route("/api/users", methods=["POST"])
@roles_required("superadmin", "aho", "director")
@rate_limit(max_calls=20, window_sec=60)
def create_user():
    d = request.json or {}
    if not d.get("email") or not d.get("name") or not d.get("password"):
        return jsonify({"error": "Заполни все поля"}), 400
    pw_raw = d["password"]
    if len(pw_raw) < 8:
        return jsonify({"error": "Пароль минимум 8 символов"}), 400
    if not any(c.isdigit() for c in pw_raw):
        return jsonify({"error": "Пароль должен содержать хотя бы одну цифру"}), 400
    if not any(c.isalpha() for c in pw_raw):
        return jsonify({"error": "Пароль должен содержать хотя бы одну букву"}), 400
    role = d.get("role", "employee")
    if role not in ROLES:
        role = next((rk for rk, rv in ROLES.items() if rv["label"] == role), "employee")
    try:
        pw = bcrypt.hashpw(pw_raw.encode(), bcrypt.gensalt()).decode()
        with get_db() as db:
            cur = db.execute(
                "INSERT INTO users (name,email,password_hash,role,department,force_password_change) VALUES (?,?,?,?,?,1)",
                (_trunc(d, "name"), d["email"].lower()[:254], pw, role, _trunc(d, "department")),
            )
            new_id = cur.lastrowid
        return jsonify({"ok": True, "id": new_id})
    except IntegrityError:
        return jsonify({"error": "Email уже используется"}), 400
    except Exception as e:
        return _db_error(e, "create_user")


@bp.route("/api/users/<int:uid>", methods=["PUT"])
@roles_required("superadmin", "aho", "director")
def update_user(uid):
    d = request.json or {}; u = request.current_user
    # Role escalation guard: cannot assign role with more privileges than your own
    _ROLE_RANK = {"employee": 0, "viewer": 0, "hr": 1, "auditor": 1, "accountant": 1,
                  "deputy": 2, "aho": 3, "director": 3, "superadmin": 4}
    my_rank = _ROLE_RANK.get(u["role"], 0)
    target_rank = _ROLE_RANK.get(d.get("role", ""), 0)
    if d.get("role") and target_rank > my_rank and u["role"] != "superadmin":
        return jsonify({"error": "Нельзя назначить роль выше своей"}), 403
    if d.get("role") and d["role"] == u["role"] and u["role"] != "superadmin":
        return jsonify({"error": "Нельзя назначить роль равную своей"}), 403
    ALLOWED = frozenset({"name", "email", "role", "active", "department", "doc_role",
                         "telegram_chat_id", "expires_at", "avatar_color"})
    sets = []; vals = []
    for k, v in d.items():
        if k in ALLOWED and k != "active":
            sets.append(f"{k}=?"); vals.append(v)
    if "active" in d:
        sets.append("active=?"); vals.append(1 if d["active"] else 0)
    if d.get("password"):
        sets.append("password_hash=?")
        vals.append(bcrypt.hashpw(d["password"].encode(), bcrypt.gensalt()).decode())
    if sets:
        try:
            with get_db() as db:
                db.execute(f"UPDATE users SET {','.join(sets)} WHERE id=?", vals + [uid])
        except IntegrityError:
            return jsonify({"error": "Email уже занят другим пользователем"}), 400
        except Exception as e:
            return _db_error(e, "update_user")
    return jsonify({"ok": True})


@bp.route("/api/users/<int:uid>", methods=["DELETE"])
@roles_required("superadmin")
def delete_user(uid):
    if uid == request.current_user["id"]: return jsonify({"error": "Нельзя удалить себя"}), 400
    with get_db() as db: db.execute("UPDATE users SET active=0 WHERE id=?", (uid,))
    return jsonify({"ok": True})


# ══════════════════════════════════════════════════════════════════════════════
#  ACTIVE / SIMPLE USERS
# ══════════════════════════════════════════════════════════════════════════════

@bp.route("/api/users/active")
def get_active_users():
    """Return list of active users for QR assign dropdown."""
    with get_db() as db:
        rows = db.execute(
            "SELECT id, name, email, role, department FROM users WHERE active=1 ORDER BY name"
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@bp.route("/api/users/simple-list")
@login_required
def get_users_simple():
    """Returns a simple list of active users for dropdowns."""
    with get_db() as db:
        rows = db.execute("SELECT id, name, department FROM users WHERE active=1 ORDER BY name").fetchall()
    return jsonify([dict(r) for r in rows])


# ══════════════════════════════════════════════════════════════════════════════
#  USER ITEMS & QR
# ══════════════════════════════════════════════════════════════════════════════

@bp.route("/api/user/<int:uid>/items")
@login_required
def get_user_items_api(uid):
    with get_db() as db:
        user = db.execute("SELECT id, name FROM users WHERE id=?", (uid,)).fetchone()
        if not user: abort(404)
        items = db.execute("""
            SELECT id, inv_num, category, model, condition, place, room,
                   purchase_price, purchase_date, serial_num
            FROM items WHERE (employee_id=? OR employee=?) AND status='Занято'
            ORDER BY category, model
        """, (uid, user["name"])).fetchall()
    return jsonify([dict(i) for i in items])


@bp.route("/api/user/<int:uid>/qr")
@login_required
def get_user_qr(uid):
    """Generates a QR code for an employee's profile."""
    return send_file(qr_png(f"{bhost()}/user/{uid}"), mimetype="image/png")


# ══════════════════════════════════════════════════════════════════════════════
#  STAFF
# ══════════════════════════════════════════════════════════════════════════════

@bp.route("/api/staff")
@login_required
def get_staff():
    """Returns all active users with their assigned equipment details."""
    with get_db() as db:
        users = db.execute("""
            SELECT u.id, u.name, u.email, u.role, u.department, u.active,
                   u.avatar_color, u.last_login,
                   COUNT(i.id) as item_count,
                   SUM(i.purchase_price) as total_value
            FROM users u
            LEFT JOIN items i ON (i.employee_id = u.id OR i.employee = u.name)
                AND i.status = 'Занято'
            WHERE u.active = 1
            GROUP BY u.id
            ORDER BY u.name
        """).fetchall()

        result = []
        for u in users:
            ud = dict(u)
            # Get this user's items
            items = db.execute("""
                SELECT id, inv_num, category, model, condition, place, room,
                       purchase_price, purchase_date, serial_num
                FROM items
                WHERE (employee_id = ? OR employee = ?) AND status = 'Занято'
                ORDER BY category, model
            """, (u["id"], u["name"])).fetchall()
            ud["items"] = [dict(i) for i in items]
            result.append(ud)

    return jsonify(result)


# ══════════════════════════════════════════════════════════════════════════════
#  DEPARTMENTS
# ══════════════════════════════════════════════════════════════════════════════

@bp.route("/api/departments")
@login_required
def get_departments():
    """Список отделов из пользователей + статистика по активам."""
    with get_db() as db:
        depts = db.execute(
            "SELECT department, COUNT(*) as user_count FROM users WHERE active=1 AND department IS NOT NULL AND department!='' GROUP BY department ORDER BY department"
        ).fetchall()
        result = []
        for d in depts:
            items = db.execute(
                """SELECT COUNT(*) as total,
                   SUM(CASE WHEN status='Занято' THEN 1 ELSE 0 END) as occupied,
                   SUM(CASE WHEN condition='Требует ремонта' THEN 1 ELSE 0 END) as repair
                   FROM items i
                   JOIN users u ON u.name = i.employee
                   WHERE u.department=? AND u.active=1""",
                (d["department"],)
            ).fetchone()
            result.append({
                "name": d["department"],
                "user_count": d["user_count"],
                "total_items": items["total"] or 0,
                "occupied": items["occupied"] or 0,
                "repair": items["repair"] or 0,
            })
    return jsonify(result)


@bp.route("/api/departments/<path:dept>/items")
@roles_required("superadmin", "aho", "hr", "auditor")
def get_dept_items(dept):
    """Все активы сотрудников отдела."""
    with get_db() as db:
        rows = db.execute(
            """SELECT i.* FROM items i
               JOIN users u ON u.name = i.employee AND u.active=1
               WHERE u.department=?
               ORDER BY u.name, i.category""",
            (dept,)
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@bp.route("/api/departments/<path:dept>/export")
@roles_required("superadmin", "aho", "auditor")
def export_dept(dept):
    """Excel-экспорт по отделу."""
    with get_db() as db:
        rows = db.execute(
            """SELECT i.inv_num, i.category, i.model, i.serial_num,
                      i.room, i.place, i.status, i.condition, i.check_date,
                      i.notes, u.name as emp_name, u.email as emp_email
               FROM items i
               JOIN users u ON u.name = i.employee AND u.active=1
               WHERE u.department=?
               ORDER BY u.name, i.category""",
            (dept,)
        ).fetchall()
    wb = Workbook(); ws = wb.active; ws.title = dept[:31]
    thin = Side(style="thin", color="CCCCCC")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    hfill = PatternFill("solid", start_color="1F4E79")
    headers = ["Инв.№", "Категория", "Модель", "Серийный №", "Кабинет", "Место",
               "Статус", "Состояние", "Дата проверки", "Примечания", "Сотрудник", "Email"]
    widths  = [10, 12, 18, 14, 12, 10, 10, 14, 12, 18, 18, 22]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10); c.fill = hfill
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = brd
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 26; ws.freeze_panes = "A2"
    ffill = PatternFill("solid", start_color="F2F2F2")
    ofill = PatternFill("solid", start_color="E2EFDA")
    for n, row in enumerate(rows, 1):
        fill = ofill if row["status"] == "Занято" else ffill
        vals = [row["inv_num"], row["category"], row["model"] or "",
                row["serial_num"] or "—", row["room"], row["place"],
                row["status"], row["condition"], row["check_date"] or "",
                row["notes"] or "", row["emp_name"], row["emp_email"] or ""]
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=n + 1, column=i, value=v)
            c.font = Font(name="Arial", size=9); c.fill = fill; c.border = brd
            c.alignment = Alignment(vertical="center")
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = dept.replace("/", "_").replace(" ", "_")
    return send_file(buf, as_attachment=True,
                     download_name=f"Dept_{fname}_{date.today()}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ══════════════════════════════════════════════════════════════════════════════
#  HISTORY
# ══════════════════════════════════════════════════════════════════════════════

@bp.route("/api/history")
@login_required
def api_history():
    limit  = min(int(request.args.get("limit", 200)), 500)  # max 500
    offset = max(int(request.args.get("offset", 0)), 0)
    action_f = request.args.get("action", "")[:100]  # limit search string
    q = "SELECT h.*, i.inv_num, i.category FROM history h LEFT JOIN items i ON h.item_id=i.id WHERE 1=1"
    params = []
    if action_f:
        q += " AND h.action LIKE ?"; params.append(f"%{action_f}%")
    q += " ORDER BY h.ts DESC LIMIT ? OFFSET ?"
    params += [limit, offset]
    with get_db() as db:
        rows = db.execute(q, params).fetchall()
    return jsonify([dict(r) for r in rows])


@bp.route("/api/history/export")
@roles_required("superadmin", "aho", "auditor")
def export_history():
    with get_db() as db:
        rows = db.execute("SELECT h.*, i.inv_num, i.category, i.model FROM history h LEFT JOIN items i ON h.item_id=i.id ORDER BY h.ts DESC").fetchall()
    wb = Workbook(); ws = wb.active; ws.title = "История изменений"
    headers = ["Дата/Время", "Сотрудник", "Инв. №", "Категория", "Действие", "Поле", "Старое значение", "Новое значение"]
    ws.append(headers)
    for row in rows:
        ws.append([row["ts"], row["user_name"], row["inv_num"] or "—", row["category"] or "—", row["action"], row["field"] or "—", str(row["old_val"] or ""), str(row["new_val"] or "")])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f"History_{date.today()}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ══════════════════════════════════════════════════════════════════════════════
#  HTML PAGES
# ══════════════════════════════════════════════════════════════════════════════

@bp.route("/admin/users")
def admin_users_page():
    with get_db() as db:
        users = db.execute("SELECT id,name,email,role,department,active FROM users ORDER BY role,name").fetchall()
    return render_template("admin_users.html", users=[dict(u) for u in users],
                           roles=ROLES, user=request.current_user)


@bp.route("/history")
def history_page():
    u = request.current_user
    with get_db() as db:
        rows = db.execute("""
            SELECT h.*, i.inv_num, i.category, i.model
            FROM history h
            LEFT JOIN items i ON h.item_id = i.id
            ORDER BY h.ts DESC LIMIT 500
        """).fetchall()
    return render_template("history.html", rows=[dict(r) for r in rows], user=u)
