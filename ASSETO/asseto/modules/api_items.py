"""Items API Blueprint — inventory, QR, assignments, issuances, write-offs, and related pages."""

import os, io, uuid, json, secrets, time, re, threading
from datetime import date, datetime, timedelta
from urllib.parse import quote

from flask import (Blueprint, render_template, request, jsonify, send_file,
                   abort, redirect, url_for, make_response)

import qrcode
from openpyxl import Workbook, load_workbook

from modules.config import (app, ROLES, CATEGORIES, PREFIXES, CONDITIONS, STATUSES,
                            UPLOADS, SIGS, ALLOWED_EXTENSIONS, SECURE_COOKIES,
                            MAX_UPLOAD_MB, JWT_EXPIRY)
from modules.db import get_db, _trunc
from modules.auth import login_required, roles_required, bhost, get_lan_ip

bp = Blueprint('items', __name__)


# ── HELPERS ──────────────────────────────────────────────────────────────────

def log_h(db, item_id, action, field=None, old_val=None, new_val=None, uid=None, uname=None):
    db.execute("INSERT INTO history (item_id,user_id,user_name,action,field,old_val,new_val) VALUES (?,?,?,?,?,?,?)",
               (item_id, uid, uname, action, field, old_val, new_val))


def _db_error(e, context=""):
    """Log full error internally, return safe generic message to client."""
    app.logger.error(f"DB error [{context}]: {e}", exc_info=True)
    return jsonify({"error": "Внутренняя ошибка сервера. Попробуйте позже."}), 500


def next_inv(cat):
    p = PREFIXES.get(cat, "ДРГ")
    with get_db() as db:
        rows = db.execute("SELECT inv_num FROM items WHERE inv_num LIKE ?", (f"{p}-%",)).fetchall()
    nums = [int(r["inv_num"].split("-")[1]) for r in rows if r["inv_num"].split("-")[1].isdigit()]
    return f"{p}-{(max(nums)+1 if nums else 1):03d}"


def qr_png(url):
    qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=8, border=2)
    qr.add_data(url); qr.make(fit=True)
    buf = io.BytesIO(); qr.make_image(fill_color="black", back_color="white").save(buf, "PNG"); buf.seek(0)
    return buf


def send_telegram(chat_id, text):
    """Отправить уведомление в Telegram. chat_id — числовой ID чата."""
    import urllib.request as _ur
    tok = os.environ.get("TELEGRAM_BOT_TOKEN", "")
    if not tok or not chat_id:
        return False
    try:
        payload = json.dumps({"chat_id": chat_id, "text": text, "parse_mode": "HTML"}).encode()
        req = _ur.Request(
            f"https://api.telegram.org/bot{tok}/sendMessage",
            data=payload, headers={"Content-Type": "application/json"}, method="POST"
        )
        _ur.urlopen(req, timeout=4)
        return True
    except Exception as e:
        app.logger.warning(f"Telegram send failed: {e}")
        return False


def _tg_notify(chat_id, text):
    """Короткий хелпер для системных уведомлений (не блокирует запрос)."""
    threading.Thread(target=send_telegram, args=(chat_id, text), daemon=True).start()


# ══════════════════════════════════════════════════════════════════════════════
#  ITEMS CRUD API
# ══════════════════════════════════════════════════════════════════════════════

@bp.route("/api/items")
def get_items():
    u = request.current_user
    q = "SELECT * FROM items WHERE 1=1"; params = []
    # Employees and viewers with no view_all see only their own items
    if not ROLES.get(u["role"], {}).get("can_view_all"):
        q += " AND (employee_id=? OR employee=?)"; params += [u["id"], u["name"]]
    else:
        for k, col in [("room", "room"), ("status", "status"), ("category", "category"), ("employee", "employee")]:
            v = request.args.get(k, "")
            if v: q += f" AND {col}=?"; params.append(v)
    q += " ORDER BY place,category"
    with get_db() as db:
        rows = db.execute(q, params).fetchall()
    return jsonify([dict(r) for r in rows])


@bp.route("/api/items", methods=["POST"])
@roles_required("superadmin", "aho")
def add_item():
    if request.is_json:
        d = request.json
    else:
        d = request.form

    u = request.current_user
    cat = d.get("category", "Другое")
    if cat not in CATEGORIES:
        return jsonify({"error": f"Недопустимая категория: {cat}"}), 400
    inv = next_inv(cat)

    emp = d.get("employee", "—")
    default_status = "Занято" if emp != "—" else "Свободно"
    status = d.get("status", default_status)
    if status not in STATUSES:
        return jsonify({"error": f"Недопустимый статус: {status}"}), 400
    condition = d.get("condition", "Хорошее")
    if condition not in CONDITIONS:
        return jsonify({"error": f"Недопустимое состояние: {condition}"}), 400

    # Validate numeric fields
    purchase_price = None
    if d.get("purchase_price"):
        try:
            purchase_price = float(d["purchase_price"])
            if purchase_price < 0: raise ValueError
        except (ValueError, TypeError):
            return jsonify({"error": "Стоимость должна быть положительным числом"}), 400

    photo_name = None
    if "photo" in request.files:
        f = request.files["photo"]
        if f.filename:
            ext = os.path.splitext(f.filename)[1].lower()
            if ext in ALLOWED_EXTENSIONS:
                photo_name = f"{uuid.uuid4().hex[:12]}{ext}"
                f.save(os.path.join(UPLOADS, photo_name))

    try:
        with get_db() as db:
            cur = db.execute(
                "INSERT INTO items (place,inv_num,category,model,serial_num,room,employee,employee_id,status,condition,check_date,notes,photo,purchase_price,purchase_date,supplier,warranty_until) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (_trunc(d, "place"), inv, cat, _trunc(d, "model"), _trunc(d, "serial_num", "—"), _trunc(d, "room"),
                 emp[:150], d.get("employee_id"), status,
                 condition, date.today().isoformat(), _trunc(d, "notes"), photo_name,
                 purchase_price, d.get("purchase_date"), d.get("supplier"), d.get("warranty_until"))
            )
            item_id = cur.lastrowid
            log_h(db, item_id, "Добавлен", uid=u["id"], uname=u["name"])
    except Exception as e:
        return _db_error(e, "add_item")

    # Telegram уведомление
    admin_chat = os.environ.get("TELEGRAM_ADMIN_CHAT_ID")
    if admin_chat:
        emp_str = f" → {emp}" if emp != "—" else ""
        _tg_notify(admin_chat,
            f"✅ <b>Добавлена техника</b>\n"
            f"📦 {d.get('category','?')} | {d.get('model','?')}\n"
            f"🏷️ Инв. №: <code>{inv}</code>{emp_str}\n"
            f"👤 Добавил: {u['name']}")

    return jsonify({"ok": True, "id": item_id, "inv_num": inv})


@bp.route("/api/items/bulk", methods=["POST"])
@roles_required("superadmin", "aho")
def add_bulk():
    d = request.json; u = request.current_user; inv_nums = []
    emp = d.get("employee", "—")
    default_status = "Занято" if emp != "—" else "Свободно"
    status = d.get("status", default_status)
    if status not in STATUSES:
        return jsonify({"error": f"Недопустимый статус: {status}"}), 400
    try:
        with get_db() as db:
            for item in d.get("items", []):
                inv = next_inv(item.get("category", "Другое"))
                cur = db.execute("INSERT INTO items (place,inv_num,category,model,serial_num,room,employee,employee_id,status,condition,check_date,notes) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                    (d.get("place", ""), inv, item.get("category", "Другое"), item.get("model", ""), item.get("serial_num", "—"),
                     d.get("room", ""), emp, d.get("employee_id"), status,
                     item.get("condition", "Хорошее"), date.today().isoformat(), item.get("notes", "")))
                log_h(db, cur.lastrowid, "Добавлен (пакетно)", uid=u["id"], uname=u["name"])
                inv_nums.append(inv)
    except Exception as e:
        return _db_error(e, "add_bulk")
    return jsonify({"ok": True, "inv_nums": inv_nums, "count": len(inv_nums)})


@bp.route("/api/items/<int:iid>", methods=["PUT"])
@login_required
def update_item(iid):
    u = request.current_user; d = request.json
    if u["role"] == "employee":
        with get_db() as db:
            item = db.execute("SELECT * FROM items WHERE id=?", (iid,)).fetchone()
        if not item or (str(item["employee_id"]) != str(u["id"]) and item["employee"] != u["name"]):
            return jsonify({"error": "Нет доступа"}), 403
        d = {k: v for k, v in d.items() if k in ("condition", "notes")}
    elif not ROLES.get(u["role"], {}).get("can_edit"):
        return jsonify({"error": "Нет доступа"}), 403
    # Validate enum fields
    if "status" in d and d["status"] not in STATUSES:
        return jsonify({"error": f"Недопустимый статус: {d['status']}"}), 400
    if "condition" in d and d["condition"] not in CONDITIONS:
        return jsonify({"error": f"Недопустимое состояние: {d['condition']}"}), 400
    if "category" in d and d["category"] not in CATEGORIES:
        return jsonify({"error": f"Недопустимая категория: {d['category']}"}), 400
    # Validate purchase_price if provided
    if "purchase_price" in d and d["purchase_price"] not in (None, "", "—"):
        try:
            d["purchase_price"] = float(d["purchase_price"])
            if d["purchase_price"] < 0: raise ValueError
        except (ValueError, TypeError):
            return jsonify({"error": "Стоимость должна быть положительным числом"}), 400
    elif "purchase_price" in d and d["purchase_price"] in ("", "—"):
        d["purchase_price"] = None

    fields = ["place", "category", "model", "serial_num", "room", "employee", "employee_id",
              "status", "condition", "notes", "purchase_price", "purchase_date", "supplier", "warranty_until"]
    with get_db() as db:
        old = dict(db.execute("SELECT * FROM items WHERE id=?", (iid,)).fetchone() or {})
        sets = [f"{f}=?" for f in fields if f in d]
        vals = [d[f] for f in fields if f in d]
        if sets: db.execute(f"UPDATE items SET {','.join(sets)} WHERE id=?", vals + [iid])
        for f in fields:
            if f in d and str(d[f]) != str(old.get(f, "")):
                log_h(db, iid, "Изменено", f, str(old.get(f, "")), str(d[f]), u["id"], u["name"])
    return jsonify({"ok": True})


@bp.route("/api/items/<int:iid>", methods=["DELETE"])
@login_required
def delete_item(iid):
    u = request.current_user
    if not ROLES[u["role"]].get("can_delete"):
        return jsonify({"error": "У вас нет прав на удаление"}), 403
    with get_db() as db:
        row = db.execute("SELECT photo FROM items WHERE id=?", (iid,)).fetchone()
        if row and row["photo"]:
            try: os.remove(os.path.join(UPLOADS, row["photo"]))
            except: pass
        db.execute("DELETE FROM items WHERE id=?", (iid,))
        db.execute("DELETE FROM history WHERE item_id=?", (iid,))
    return jsonify({"ok": True})


@bp.route("/api/items/<int:iid>/photo", methods=["POST", "DELETE"])
@login_required
def photo(iid):
    u = request.current_user
    if request.method == "DELETE":
        if u["role"] not in ("superadmin", "aho"): return jsonify({"error": "Нет доступа"}), 403
        with get_db() as db:
            old = db.execute("SELECT photo FROM items WHERE id=?", (iid,)).fetchone()
            if old and old["photo"]:
                try: os.remove(os.path.join(UPLOADS, old["photo"]))
                except: pass
            db.execute("UPDATE items SET photo=NULL WHERE id=?", (iid,))
            log_h(db, iid, "Фото удалено", uid=u["id"], uname=u["name"])
        return jsonify({"ok": True})
    if "photo" not in request.files: return jsonify({"ok": False})
    f = request.files["photo"]
    ext = os.path.splitext(f.filename)[1].lower() if f.filename else '.jpg'
    if ext not in ALLOWED_EXTENSIONS:
        return jsonify({"error": "Недопустимый формат файла. Разрешены: JPG, PNG, GIF, WEBP"}), 400
    name = f"{iid}_{uuid.uuid4().hex[:8]}{ext}"
    with get_db() as db:
        old = db.execute("SELECT photo,employee_id,employee FROM items WHERE id=?", (iid,)).fetchone()
        if u["role"] == "employee":
            if not old or (str(old["employee_id"]) != str(u["id"]) and old["employee"] != u["name"]):
                return jsonify({"error": "Нет доступа"}), 403
        if old and old["photo"]:
            try: os.remove(os.path.join(UPLOADS, old["photo"]))
            except: pass
    f.save(os.path.join(UPLOADS, name))
    with get_db() as db:
        db.execute("UPDATE items SET photo=? WHERE id=?", (name, iid))
        log_h(db, iid, "Фото обновлено", uid=u["id"], uname=u["name"])
    return jsonify({"ok": True, "photo": name})


@bp.route("/api/items/<int:iid>/verify", methods=["POST"])
@login_required
def verify_item(iid):
    u = request.current_user
    if "photo" not in request.files:
        return jsonify({"error": "Нужно фото подтверждение"}), 400
    f = request.files["photo"]
    ext = os.path.splitext(f.filename)[1].lower() if f.filename else '.jpg'
    if ext not in ALLOWED_EXTENSIONS:
        return jsonify({"error": "Недопустимый формат файла"}), 400
    name = f"verify_{iid}_{uuid.uuid4().hex[:8]}{ext}"

    with get_db() as db:
        item = db.execute("SELECT * FROM items WHERE id=?", (iid,)).fetchone()
        if not item: return jsonify({"error": "Ни найдено"}), 404

        # Check permissions
        is_owner = str(item["employee_id"]) == str(u["id"]) or item["employee"] == u["name"]
        can_edit = ROLES[u["role"]]["can_edit"]
        if not is_owner and not can_edit:
            return jsonify({"error": "Нет доступа к верификации этого оборудования"}), 403

        # Save photo
        f.save(os.path.join(UPLOADS, name))

        # Update item
        db.execute("UPDATE items SET photo=?, check_date=? WHERE id=?", (name, date.today().isoformat(), iid))
        log_h(db, iid, "Верификация (фото-подтверждение)", uid=u["id"], uname=u["name"])

    return jsonify({"ok": True, "photo": name})


@bp.route("/api/items/<int:iid>/history")
@login_required
def item_history(iid):
    """Returns history for a specific item."""
    with get_db() as db:
        rows = db.execute("""
            SELECT * FROM history WHERE item_id=? ORDER BY ts DESC LIMIT 50
        """, (iid,)).fetchall()
    return jsonify([dict(r) for r in rows])


@bp.route("/api/items/<int:iid>/valuation")
@login_required
def item_valuation(iid):
    """Рассчитать остаточную стоимость актива (линейная амортизация)."""
    with get_db() as db:
        item = db.execute("SELECT * FROM items WHERE id=?", (iid,)).fetchone()
    if not item:
        return jsonify({"error": "Не найдено"}), 404
    item = dict(item)
    if not item.get("purchase_price") or not item.get("purchase_date"):
        return jsonify({"purchase_price": None, "residual_value": None,
                        "depreciation_pct": None, "message": "Нет данных о стоимости"})
    # Default useful life by category (years)
    useful_life = {
        "Ноутбук": 3, "Монитор": 5, "Кресло": 7, "Стол": 10,
        "Принтер": 4, "Телефон": 2, "Клавиатура": 3, "Мышь": 2,
    }.get(item.get("category", ""), 5)
    from datetime import date as _d
    today = _d.today()
    try:
        purchased = _d.fromisoformat(item["purchase_date"])
    except:
        return jsonify({"error": "Неверный формат даты покупки"}), 400
    years_used = (today - purchased).days / 365.25
    depreciation_pct = min(100.0, (years_used / useful_life) * 100)
    residual_value = round(item["purchase_price"] * (1 - depreciation_pct / 100), 2)
    return jsonify({
        "purchase_price": item["purchase_price"],
        "purchase_date": item["purchase_date"],
        "years_used": round(years_used, 1),
        "useful_life_years": useful_life,
        "depreciation_pct": round(depreciation_pct, 1),
        "residual_value": max(0, residual_value),
        "category": item.get("category")
    })


# ══════════════════════════════════════════════════════════════════════════════
#  REASSIGN, MAINTENANCE, WRITE-OFF
# ══════════════════════════════════════════════════════════════════════════════

@bp.route("/api/items/<inv_num>/reassign", methods=["POST"])
@roles_required("superadmin", "aho")
def reassign_item(inv_num):
    """Quickly reassign an asset to another employee."""
    d = request.json
    uid = d.get("user_id")
    if not uid: return jsonify({"error": "Выберите сотрудника"}), 400

    with get_db() as db:
        # Get target user
        user = db.execute("SELECT id, name FROM users WHERE id=?", (uid,)).fetchone()
        if not user: return jsonify({"error": "Сотрудник не найден"}), 404

        # Get current item info
        item = db.execute("SELECT id, employee FROM items WHERE inv_num=?", (inv_num,)).fetchone()
        if not item: return jsonify({"error": "Предмет не найден"}), 404

        old_val = item["employee"]
        new_val = user["name"]

        # Update item
        db.execute("UPDATE items SET employee=?, employee_id=?, status='Занято' WHERE inv_num=?",
                   (new_val, uid, inv_num))

        # Log history
        db.execute("INSERT INTO history (item_id, user_id, user_name, action, field, old_val, new_val) VALUES (?,?,?,?,?,?,?)",
                   (item["id"], request.current_user["id"], request.current_user["name"],
                    "Переназначение", "employee", old_val, new_val))

    # Telegram уведомление
    admin_chat = os.environ.get("TELEGRAM_ADMIN_CHAT_ID")
    if admin_chat:
        _tg_notify(admin_chat,
            f"🔄 <b>Переназначение техники</b>\n"
            f"🏷️ Инв. №: <code>{inv_num}</code>\n"
            f"👤 Кому: {new_val}\n"
            f"📋 Было: {old_val or '—'}\n"
            f"✍️ Назначил: {request.current_user['name']}")

    return jsonify({"ok": True, "new_owner": new_val})


@bp.route("/api/items/<inv_num>/maintenance", methods=["POST"])
@login_required
def report_maintenance(inv_num):
    """Report an issue with an asset."""
    d = request.json
    desc = d.get("description")
    if not desc: return jsonify({"error": "Опишите проблему"}), 400

    with get_db() as db:
        item = db.execute("SELECT id FROM items WHERE inv_num=?", (inv_num,)).fetchone()
        if not item: return jsonify({"error": "Предмет не найден"}), 404

        db.execute("""INSERT INTO maintenance (item_id, reported_by_id, reported_by_name, description)
                      VALUES (?,?,?,?)""",
                   (item["id"], request.current_user["id"], request.current_user["name"], desc))

        # Update item condition
        db.execute("UPDATE items SET condition='Требует ремонта' WHERE inv_num=?", (inv_num,))

    # Telegram уведомление АХО/admin
    admin_chat = os.environ.get("TELEGRAM_ADMIN_CHAT_ID")
    if admin_chat:
        _tg_notify(admin_chat,
            f"🔧 <b>Заявка на ремонт</b>\n"
            f"🏷️ Инв. №: <code>{inv_num}</code>\n"
            f"📝 Описание: {desc[:200]}\n"
            f"👤 Заявитель: {request.current_user['name']}")

    return jsonify({"ok": True})


@bp.route("/api/items/<path:inv_num>/writeoff", methods=["POST"])
@roles_required("superadmin", "aho", "director")
def writeoff_item(inv_num):
    """Write off an asset: sets condition=Списано, clears employee, logs with act details."""
    d = request.json or {}
    reason = (d.get("reason") or "").strip()
    if not reason:
        return jsonify({"error": "Укажите причину списания"}), 400
    act_num = (d.get("act_num") or "").strip()
    authorized_by = (d.get("authorized_by") or "").strip()
    u = request.current_user
    with get_db() as db:
        item = db.execute("SELECT * FROM items WHERE inv_num=?", (inv_num,)).fetchone()
        if not item: return jsonify({"error": "Актив не найден"}), 404
        if item["condition"] == "Списано":
            return jsonify({"error": "Актив уже списан"}), 400
        old_emp = item["employee"] or "—"
        db.execute(
            "UPDATE items SET condition='Списано', status='Свободно', employee='—', employee_id=NULL WHERE inv_num=?",
            (inv_num,)
        )
        note = f"Акт: {act_num or '—'} | Причина: {reason} | Утверждено: {authorized_by or '—'}"
        log_h(db, item["id"], f"Списание", "condition", item["condition"], "Списано", u["id"], u["name"])
        log_h(db, item["id"], note, uid=u["id"], uname=u["name"])
        # Store writeoff details in notes for act printing
        existing_notes = item["notes"] or ""
        wo_stamp = f"[СПИСАНИЕ {date.today().isoformat()}] Акт: {act_num or '—'}, Причина: {reason}, Утв.: {authorized_by or '—'}, Исполнитель: {u['name']}"
        db.execute("UPDATE items SET notes=? WHERE inv_num=?",
                   ((wo_stamp + "\n" + existing_notes).strip()[:2000], inv_num))
    return jsonify({"ok": True})


# ══════════════════════════════════════════════════════════════════════════════
#  ASSIGN / UNASSIGN / TRANSFER
# ══════════════════════════════════════════════════════════════════════════════

@bp.route("/api/items/<int:item_id>/assign", methods=["POST"])
@login_required
def quick_assign_item(item_id):
    r = request.current_user["role"]
    if not (ROLES.get(r, {}).get("can_issue") or ROLES.get(r, {}).get("can_edit")):
        return jsonify({"error": "Нет доступа"}), 403
    data = request.get_json() or {}
    user_id = data.get("user_id")
    if not user_id:
        return jsonify({"error": "user_id required"}), 400
    with get_db() as db:
        item = db.execute("SELECT id, inv_num, model, status FROM items WHERE id=?", (item_id,)).fetchone()
        if not item: return jsonify({"error": "Техника не найдена"}), 404
        if item["status"] == "Занято": return jsonify({"error": "Уже занято"}), 409
        user = db.execute("SELECT id, name FROM users WHERE id=?", (user_id,)).fetchone()
        if not user: return jsonify({"error": "Пользователь не найден"}), 404
        db.execute(
            "UPDATE items SET employee_id=?, employee=?, status='Занято', updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (user["id"], user["name"], item_id)
        )
        db.commit()
    return jsonify({"ok": True, "inv_num": item["inv_num"], "employee": user["name"]})


@bp.route("/api/items/<int:item_id>/unassign", methods=["POST"])
@login_required
def quick_unassign_item(item_id):
    r = request.current_user["role"]
    if not (ROLES.get(r, {}).get("can_issue") or ROLES.get(r, {}).get("can_edit")):
        return jsonify({"error": "Нет доступа"}), 403
    with get_db() as db:
        db.execute(
            "UPDATE items SET employee_id=NULL, employee=NULL, status='Свободно', updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (item_id,)
        )
        db.commit()
    return jsonify({"ok": True})


@bp.route("/api/items/<int:iid>/assign", methods=["POST"])
@roles_required("superadmin", "aho", "hr")
def assign_item(iid):
    """Quickly assign or release an item via QR-scan page."""
    d = request.json or {}
    u = request.current_user
    emp_id   = d.get("employee_id")   # None → release
    emp_name = d.get("employee_name", "")

    with get_db() as db:
        item = db.execute("SELECT * FROM items WHERE id=?", (iid,)).fetchone()
        if not item:
            return jsonify({"error": "Предмет не найден"}), 404

        old_emp = item["employee"] or "—"

        if emp_id:
            # Assign to employee
            emp = db.execute("SELECT * FROM users WHERE id=? AND active=1", (emp_id,)).fetchone()
            if not emp:
                return jsonify({"error": "Сотрудник не найден"}), 404
            emp_name = emp["name"]
            db.execute(
                "UPDATE items SET employee_id=?, employee=?, status='Занято' WHERE id=?",
                (emp_id, emp_name, iid)
            )
            action = f"QR-назначение: {old_emp} → {emp_name}"
        else:
            # Release item
            db.execute(
                "UPDATE items SET employee_id=NULL, employee='—', status='Свободно' WHERE id=?",
                (iid,)
            )
            action = f"QR-освобождение: {old_emp} → Свободно"

        log_h(db, iid, action, uid=u["id"], uname=u["name"])

    return jsonify({"ok": True, "action": action,
                    "employee": emp_name if emp_id else "—",
                    "status": "Занято" if emp_id else "Свободно"})


@bp.route("/api/items/<int:iid>/transfer", methods=["POST"])
@roles_required("superadmin", "aho", "hr")
def transfer_item(iid):
    """Переназначить актив другому сотруднику с записью в историю."""
    d = request.json or {}
    u = request.current_user
    new_emp_id   = d.get("employee_id")
    note         = (d.get("note") or "").strip()[:500]

    with get_db() as db:
        item = db.execute("SELECT * FROM items WHERE id=?", (iid,)).fetchone()
        if not item:
            return jsonify({"error": "Актив не найден"}), 404
        old_emp = item["employee"] or "—"

        if new_emp_id:
            emp = db.execute("SELECT * FROM users WHERE id=? AND active=1", (new_emp_id,)).fetchone()
            if not emp:
                return jsonify({"error": "Сотрудник не найден"}), 404
            new_name = emp["name"]
            db.execute(
                "UPDATE items SET employee_id=?, employee=?, status='Занято' WHERE id=?",
                (new_emp_id, new_name, iid)
            )
            action = f"Передача: {old_emp} → {new_name}"
        else:
            new_name = "—"
            db.execute(
                "UPDATE items SET employee_id=NULL, employee='—', status='Свободно' WHERE id=?",
                (iid,)
            )
            action = f"Освобождение: {old_emp} → Свободно"

        if note:
            action += f" | {note}"
        log_h(db, iid, action, "employee", old_emp, new_name, u["id"], u["name"])

    return jsonify({"ok": True, "action": action})


# ══════════════════════════════════════════════════════════════════════════════
#  FREE ITEMS, IMPORT, BULK, AUDIT
# ══════════════════════════════════════════════════════════════════════════════

@bp.route("/api/items/free")
@login_required
def get_free_items_api():
    with get_db() as db:
        items = db.execute("""
            SELECT id, inv_num, category, model, condition, purchase_price, room
            FROM items
            WHERE status IN ('Свободно','На складе','Free','В наличии','Свободна')
            ORDER BY category, model
        """).fetchall()
    return jsonify([dict(i) for i in items])


@bp.route("/api/items/import", methods=["POST"])
@roles_required("superadmin", "aho")
def import_excel():
    if "file" not in request.files:
        return jsonify({"error": "Нет файла"}), 400
    f = request.files["file"]
    if not f.filename.endswith((".xlsx", ".xls")):
        return jsonify({"error": "Только .xlsx файлы"}), 400
    try:
        wb = load_workbook(io.BytesIO(f.read()), data_only=True)
        ws = wb.active
        u = request.current_user
        imported = 0; errors = []
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        # Map columns: place, room, category, model, serial_num, employee, condition, notes
        col = {}
        mapping = {
            "место": "place", "кабинет": "room", "наименование": "category",
            "категория": "category", "модель": "model", "серийный": "serial_num",
            "сотрудник": "employee", "состояние": "condition", "примечания": "notes"
        }
        for i, h in enumerate(headers):
            for key, field in mapping.items():
                if key in h: col[field] = i; break
        required = {"place", "room", "category"}
        missing = required - set(col.keys())
        if missing:
            return jsonify({"error": f"Не найдены колонки: {', '.join(missing)}. Нужны: Место, Кабинет, Наименование"}), 400
        with get_db() as db:
            for ri, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                try:
                    def g(field, default=""):
                        idx = col.get(field)
                        val = row[idx] if idx is not None and idx < len(row) else None
                        return str(val).strip() if val is not None else default
                    place = g("place"); room = g("room"); cat = g("category", "Другое")
                    if not place or not room: continue
                    if cat not in CATEGORIES: cat = "Другое"
                    inv = next_inv(cat)
                    emp = g("employee", "—") or "—"
                    cond = g("condition", "Хорошее")
                    if cond not in CONDITIONS: cond = "Хорошее"
                    cur = db.execute(
                        "INSERT INTO items (place,inv_num,category,model,serial_num,room,employee,status,condition,check_date,notes) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                        (place, inv, cat, g("model"), g("serial_num", "—"), room, emp,
                         "Занято" if emp != "—" else "Свободно", cond, date.today().isoformat(), g("notes"))
                    )
                    log_h(db, cur.lastrowid, "Импорт из Excel", uid=u["id"], uname=u["name"])
                    imported += 1
                except Exception as e:
                    errors.append(f"Строка {ri}: {e}")
        return jsonify({"ok": True, "imported": imported, "errors": errors})
    except Exception as e:
        return jsonify({"error": f"Ошибка чтения файла: {e}"}), 400


@bp.route("/api/items/bulk-update", methods=["POST"])
@roles_required("superadmin", "aho")
def bulk_update_items():
    d   = request.json or {}; u = request.current_user
    ids = d.get("ids", [])
    if not ids: return jsonify({"error": "Нет ID"}), 400
    fields = {k: v for k, v in d.items() if k in ("status", "condition", "room", "employee", "employee_id") and v is not None}
    if not fields: return jsonify({"error": "Нет полей"}), 400
    sets = ", ".join(f"{k}=?" for k in fields); vals = list(fields.values())
    with get_db() as db:
        for iid in ids:
            db.execute(f"UPDATE items SET {sets} WHERE id=?", vals + [iid])
            log_h(db, iid, "Массовое обновление", uid=u["id"], uname=u["name"])
    return jsonify({"ok": True, "updated": len(ids)})


@bp.route("/api/items/bulk-delete", methods=["POST"])
@login_required
def bulk_delete_items():
    u = request.current_user
    if not ROLES[u["role"]].get("can_delete"):
        return jsonify({"error": "Нет прав на массовое удаление"}), 403
    d   = request.json or {}; ids = d.get("ids", [])
    if not ids: return jsonify({"error": "Нет ID"}), 400
    with get_db() as db:
        for iid in ids:
            row = db.execute("SELECT photo FROM items WHERE id=?", (iid,)).fetchone()
            if row and row["photo"]:
                try: os.remove(os.path.join(UPLOADS, row["photo"]))
                except: pass
            db.execute("DELETE FROM items WHERE id=?", (iid,))
            db.execute("DELETE FROM history WHERE item_id=?", (iid,))
    return jsonify({"ok": True, "deleted": len(ids)})


@bp.route("/api/items/audit", methods=["POST"])
@roles_required("superadmin", "aho", "auditor")
def audit_items():
    d = request.json or {}; u = request.current_user; ids = d.get("ids", [])
    if not ids: return jsonify({"error": "Нет ID"}), 400
    with get_db() as db:
        for iid in ids:
            db.execute("UPDATE items SET check_date=? WHERE id=?", (date.today().isoformat(), iid))
            log_h(db, iid, "Аудит проведён", uid=u["id"], uname=u["name"])
    return jsonify({"ok": True, "audited": len(ids)})


# ══════════════════════════════════════════════════════════════════════════════
#  INVENTORY SESSIONS API
# ══════════════════════════════════════════════════════════════════════════════

@bp.route("/api/inventory/sessions")
def list_inv_sessions():
    with get_db() as db:
        rows = db.execute(
            "SELECT * FROM inventory_sessions ORDER BY created_at DESC LIMIT 50"
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@bp.route("/api/inventory/sessions", methods=["POST"])
@roles_required("superadmin", "aho")
def create_inv_session():
    u = request.current_user
    d = request.json or {}
    title = (d.get("title") or "").strip()
    if not title:
        return jsonify({"error": "Укажите название"}), 400
    dept = d.get("department", "")
    with get_db() as db:
        # Collect items to check
        if dept:
            items = db.execute(
                "SELECT i.id FROM items i JOIN users u ON u.name=i.employee WHERE u.department=?", (dept,)
            ).fetchall()
        else:
            items = db.execute("SELECT id FROM items WHERE status='Занято'").fetchall()
        cur = db.execute(
            "INSERT INTO inventory_sessions (title,created_by_id,created_by_name,department,total_items) VALUES (?,?,?,?,?)",
            (title, u["id"], u["name"], dept, len(items))
        )
        sid = cur.lastrowid
        for item in items:
            db.execute(
                "INSERT INTO inventory_checks (session_id,item_id) VALUES (?,?)",
                (sid, item["id"])
            )
    return jsonify({"ok": True, "session_id": sid, "total_items": len(items)})


@bp.route("/api/inventory/sessions/<int:sid>", methods=["GET"])
@login_required
def get_inv_session(sid):
    with get_db() as db:
        session = db.execute("SELECT * FROM inventory_sessions WHERE id=?", (sid,)).fetchone()
        if not session:
            return jsonify({"error": "Не найдено"}), 404
        checks = db.execute(
            """SELECT c.*,i.inv_num,i.category,i.model,i.room,i.employee
               FROM inventory_checks c
               JOIN items i ON c.item_id=i.id
               WHERE c.session_id=? ORDER BY i.room,i.category""",
            (sid,)
        ).fetchall()
    return jsonify({"session": dict(session), "checks": [dict(r) for r in checks]})


@bp.route("/api/inventory/check/<int:cid>", methods=["POST"])
@roles_required("superadmin", "aho", "auditor")
def submit_inv_check(cid):
    """Сотрудник/АХО подтверждает наличие актива."""
    u = request.current_user
    d = request.json or {}
    status = d.get("status", "found")  # found | not_found | damaged
    note   = (d.get("note") or "").strip()[:500]
    with get_db() as db:
        check = db.execute("SELECT * FROM inventory_checks WHERE id=?", (cid,)).fetchone()
        if not check:
            return jsonify({"error": "Не найдено"}), 404
        db.execute(
            """UPDATE inventory_checks SET status=?,checked_by_id=?,checked_by_name=?,
               note=?,checked_at=CURRENT_TIMESTAMP WHERE id=?""",
            (status, u["id"], u["name"], note, cid)
        )
        # Update session progress
        session_id = check["session_id"]
        checked = db.execute(
            "SELECT COUNT(*) FROM inventory_checks WHERE session_id=? AND status!='pending'",
            (session_id,)
        ).fetchone()[0]
        total = db.execute(
            "SELECT total_items FROM inventory_sessions WHERE id=?", (session_id,)
        ).fetchone()[0]
        db.execute(
            "UPDATE inventory_sessions SET checked_items=? WHERE id=?",
            (checked, session_id)
        )
        if checked >= total:
            db.execute(
                "UPDATE inventory_sessions SET status='completed',completed_at=CURRENT_TIMESTAMP WHERE id=?",
                (session_id,)
            )
    return jsonify({"ok": True, "progress": f"{checked}/{total}"})


# ══════════════════════════════════════════════════════════════════════════════
#  ISSUANCES
# ══════════════════════════════════════════════════════════════════════════════

@bp.route("/api/issuances")
def list_issuances():
    with get_db() as db:
        rows = db.execute("SELECT * FROM issuances ORDER BY created_at DESC LIMIT 100").fetchall()
    return jsonify([dict(r) for r in rows])


@bp.route("/api/issuances", methods=["POST"])
@roles_required("superadmin", "aho", "hr")
def create_issuance():
    d = request.json; u = request.current_user
    emp_id = d.get("employee_id"); emp_name = d.get("employee_name", ""); item_ids = d.get("item_ids", [])
    if not emp_id or not item_ids: return jsonify({"error": "Нужен сотрудник и список техники"}), 400
    with get_db() as db:
        for iid in item_ids:
            db.execute("UPDATE items SET employee_id=?,employee=?,status='Занято' WHERE id=?", (emp_id, emp_name, iid))
            log_h(db, iid, f"Выдано: {emp_name}", uid=u["id"], uname=u["name"])
        cur = db.execute("INSERT INTO issuances (employee_id,employee_name,issued_by,issued_by_name,items_json) VALUES (?,?,?,?,?)",
            (emp_id, emp_name, u["id"], u["name"], json.dumps(item_ids)))
    return jsonify({"ok": True, "issuance_id": cur.lastrowid})


@bp.route("/api/issuances/<int:iid>/confirm", methods=["POST"])
@login_required
def confirm_issuance(iid):
    u = request.current_user
    with get_db() as db:
        iss = db.execute("SELECT * FROM issuances WHERE id=?", (iid,)).fetchone()
        if not iss: return jsonify({"error": "Не найдено"}), 404
        if u["role"] == "employee" and iss["employee_id"] != u["id"]:
            return jsonify({"error": "Нет доступа"}), 403
        db.execute("UPDATE issuances SET status='confirmed',confirmed_at=CURRENT_TIMESTAMP WHERE id=?", (iid,))
    return jsonify({"ok": True})


# ══════════════════════════════════════════════════════════════════════════════
#  QR API
# ══════════════════════════════════════════════════════════════════════════════

@bp.route("/api/qr/<path:inv_num>")
@login_required
def get_item_qr(inv_num):
    """Generates a QR code for an item."""
    try:
        url = f"{bhost()}/asset/{inv_num}"
        img_buf = qr_png(url)
        return send_file(img_buf, mimetype="image/png", max_age=0)
    except Exception as e:
        app.logger.error(f"QR Error: {e}")
        return abort(500)


@bp.route("/api/qr_employee/<path:enc_name>")
@login_required
def get_employee_qr(enc_name):
    """Generates a QR code for an employee name string."""
    return send_file(qr_png(f"{bhost()}/employee/{enc_name}"), mimetype="image/png")


# ══════════════════════════════════════════════════════════════════════════════
#  HTML PAGES
# ══════════════════════════════════════════════════════════════════════════════

@bp.route("/inventory-session")
@login_required
def inventory_sessions_page():
    u = request.current_user
    return render_template("inventory_sessions.html", user=u, current_user=u,
        role_info=ROLES.get(u["role"], {}), roles=ROLES)


@bp.route("/equipment")
@login_required
def equipment_page():
    u = request.current_user
    return render_template("equipment.html", user=u, current_user=u,
        role_info=ROLES.get(u["role"], {}), roles=ROLES,
        categories=CATEGORIES, prefixes=PREFIXES)


@bp.route("/inventory")
@login_required
def inventory_page():
    u = request.current_user
    return render_template("inventory.html", user=u, current_user=u,
        role_info=ROLES.get(u["role"], {}), roles=ROLES)


@bp.route("/inventory/<int:sid>")
@login_required
def inventory_session_page(sid):
    u = request.current_user
    with get_db() as db:
        session = db.execute("SELECT * FROM inventory_sessions WHERE id=?", (sid,)).fetchone()
    if not session: abort(404)
    return render_template("inventory_session.html", session=dict(session), user=u, current_user=u,
        role_info=ROLES.get(u["role"], {}), roles=ROLES)


@bp.route("/asset/<inv_num>/print")
@login_required
def print_label_page(inv_num):
    with get_db() as db:
        item = db.execute("SELECT * FROM items WHERE inv_num=?", (inv_num,)).fetchone()
    if not item: abort(404)
    return render_template("print_label.html", item=dict(item), host=bhost())


@bp.route("/asset/<path:inv_num>/writeoff-act")
@roles_required("superadmin", "aho", "director", "accountant", "auditor")
def writeoff_act_page(inv_num):
    """Printable write-off act for a written-off asset."""
    with get_db() as db:
        item = db.execute("SELECT * FROM items WHERE inv_num=?", (inv_num,)).fetchone()
    if not item: abort(404)
    item = dict(item)
    # Parse write-off metadata from notes
    wo_info = {"act_num": "—", "reason": "—", "authorized_by": "—", "date": "—", "executor": "—"}
    notes = item.get("notes") or ""
    m = re.search(r'\[СПИСАНИЕ (\d{4}-\d{2}-\d{2})\] Акт: (.+?), Причина: (.+?), Утв\.: (.+?), Исполнитель: (.+?)$', notes, re.MULTILINE)
    if m:
        wo_info = {"date": m.group(1), "act_num": m.group(2), "reason": m.group(3),
                   "authorized_by": m.group(4), "executor": m.group(5)}
    return render_template("writeoff_act.html", item=item, wo=wo_info, user=request.current_user, host=bhost())


@bp.route("/qr-sheet")
@login_required
def qr_sheet():
    with get_db() as db:
        items = db.execute("SELECT id,inv_num,category,model,room,employee,place FROM items ORDER BY room,inv_num").fetchall()
    return render_template("qr_sheet.html", items=[dict(i) for i in items])


@bp.route("/qr-sheet-employees")
@login_required
def qr_sheet_employees():
    with get_db() as db:
        rows = db.execute("SELECT employee,COUNT(*) as cnt FROM items WHERE employee IS NOT NULL AND employee!='' AND employee!='—' GROUP BY employee ORDER BY employee").fetchall()
    emps = [{"name": r["employee"], "name_encoded": quote(r["employee"]), "count": r["cnt"]} for r in rows]
    return render_template("qr_sheet_employees.html", employees=emps)


@bp.route("/qr-print")
@login_required
def qr_print_page():
    """Advanced QR print page: filter by room, category, employee, department. Thermal printer support."""
    with get_db() as db:
        items = db.execute("""
            SELECT i.id, i.inv_num, i.category, i.model, i.room, i.employee,
                   i.place, i.serial_num, i.condition,
                   u.department
            FROM items i
            LEFT JOIN users u ON u.name = i.employee AND i.employee != '—'
            ORDER BY i.room, i.category, i.inv_num
        """).fetchall()
        rooms  = [r[0] for r in db.execute("SELECT DISTINCT room FROM items WHERE room IS NOT NULL ORDER BY room").fetchall()]
        cats   = [r[0] for r in db.execute("SELECT DISTINCT category FROM items ORDER BY category").fetchall()]
        emps   = [r[0] for r in db.execute("SELECT DISTINCT employee FROM items WHERE employee IS NOT NULL AND employee != '—' ORDER BY employee").fetchall()]
        depts  = [r[0] for r in db.execute("SELECT DISTINCT department FROM users WHERE department IS NOT NULL AND department != '' ORDER BY department").fetchall()]
    u = request.current_user
    return render_template("qr_print.html",
        items=[dict(i) for i in items],
        rooms=rooms, categories=cats, employees=emps, departments=depts,
        user=u, role_info=ROLES.get(u["role"], {}), host=bhost())


@bp.route("/employee/<int:uid>/print")
@login_required
def print_employee_label(uid):
    """Page for printing an employee's QR badge."""
    with get_db() as db:
        user = db.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
    if not user: abort(404)
    return render_template("print_employee.html", user=dict(user), host=bhost())


@bp.route("/issuances/<int:iid>/print-act")
@roles_required("superadmin", "aho", "hr", "director")
def print_issuance_act(iid):
    """Printable transfer act (Акт приёма-передачи) for an issuance."""
    with get_db() as db:
        iss = db.execute("SELECT * FROM issuances WHERE id=?", (iid,)).fetchone()
        if not iss: abort(404)
        iss = dict(iss)
        item_ids = json.loads(iss.get("items_json") or "[]")
        items = []
        for item_id in item_ids:
            row = db.execute("SELECT * FROM items WHERE id=?", (item_id,)).fetchone()
            if row: items.append(dict(row))
        emp = db.execute("SELECT * FROM users WHERE id=?", (iss["employee_id"],)).fetchone()
        emp = dict(emp) if emp else {}
    return render_template("transfer_act.html", iss=iss, items=items, emp=emp,
                           user=request.current_user, host=bhost())
